# ##################################### Final Script to get exceptions that are about to expire in next one month ###################################################################################

import openpyxl
import json
import os
import re
import datetime
from dateutil.relativedelta import *

def has_expired(dt):
    return dt < datetime.datetime.now()

def prWarn(msg): print(f"\033[93mWarning: {msg}\033[00m")

def parse_date(dt,format=None):        
    if isinstance(dt, datetime.datetime):
        if format:
            return dt.strftime(format)
        else:
            return dt
    elif dt == "Permanent":
        return datetime.datetime.strptime(f"1/1/9999", "%d/%m/%Y")
    reg_exp = re.search("^Q(\d).(\d{4})$", dt)
    if reg_exp:
        quarter = int(reg_exp.group(1))
        year = int(reg_exp.group(2))
        if quarter == 4:
            year += 1
            month = 1
        else:
            month = quarter*3+1
        dt = datetime.datetime.strptime(f"1/{month}/{year}", "%d/%m/%Y")
        if format:
            return dt.strftime(format)
        return dt
    prWarn(f"Unable to parse date: {dt}")
    return datetime.datetime.strptime(f"1/1/1970", "%d/%m/%Y")

def parse_cloud_id(cloud_id):
    cloud_id = str(cloud_id).strip()
    if re.search("^\d{1,12}$",cloud_id):
        return cloud_id.zfill(12) 
    return cloud_id

class Excel:
    def __init__(self, path):
        if not os.path.isfile(path):
            print(f"{path} does not exist")
            exit(1)
        self.path = path
        self.workbook = openpyxl.load_workbook(f"{path}")
        self.sheet = self.workbook["Prisma Security Exception"]
        self.col_mapping = {
            "account_id": 2,
            "sgs_ticket_id": 7,
            "validity": 9,
            "status": 24,
            "L7": 27,
            "L6": 28,
            "L5": 29,
            "L4": 30, 
        }
        self.valid_control_ids = {}
        self.valid_controls = {}

    def get_exceptions_by_ticket(self, sep):
        L7_level = None
        L6_level = None
        L5_level = None
        L4_level = None
        accounts = set()
        exception_id = None
        expiration_date = None
        for row in self.sheet[2:self.sheet.max_row]:
            for key in self.col_mapping.keys():
                exec(f'global {key}; {key} = row[self.col_mapping["{key}"]].value')
            if validity == 45307: # This is one of the Invalid date from the sheet that is not convered by the parse_date function
                continue
            elif validity == 45111: # This is one of the Invalid date from the sheet that is not convered by the parse_date function
                continue
            # print(validity)
            # print(parse_date(validity))
            expiry_date = parse_date(validity)
            todays_date = datetime.datetime.now()
            one_month_later_date = todays_date + relativedelta(months=+1)
            if str(status).lower() == "true" and sgs_ticket_id == sep and expiry_date < one_month_later_date and not has_expired(expiry_date): 
                exception_id = sgs_ticket_id
                accounts.add(str(parse_cloud_id(account_id)))
                expiration_date = parse_date(validity,"%Y-%m-%d")
                L7_level = L7
                L6_level = L6
                L5_level = L5
                L4_level = L4
        return {'ticket_id': exception_id, 'expiry_date': expiration_date, 'L7_board_area': L7_level, 'L6_board_area': L6_level, 'L5_board_area': L5_level, 'L4_board_area': L4_level, 'accounts': len(accounts)}
 
path = os.path.join(os.getcwd(),"datamodel","data","SGS_Excel","Chef-and-Secure-by-Default-exceptions_8-14.xlsx")

ticket_ids = set()
exception_workbook = openpyxl.load_workbook(f"{path}")
exceptions_sheet = exception_workbook["Prisma Security Exception"]

for column in exceptions_sheet.iter_cols():
    column_name = column[0].value
    if column_name == 'SGS Ticket ID':
        for i, cell in enumerate(column):
            if i > 0:
              ticket_ids.add(cell.value)
uniq_ticket_ids = list(ticket_ids)

exceptions_list = []

if __name__ == "__main__":
    if not os.path.isfile(path):
        print(f'Copy Orca_exceptions.xlsx to {os.path.join(os.getcwd(),"datamodel","data","SGS_Excel")}')
        exit(1)
    excel = Excel(path)
    for ticket in uniq_ticket_ids:
      exception = excel.get_exceptions_by_ticket(ticket)
      if exception.get('ticket_id', None):
        exceptions_list.append(exception)

today = datetime.date.today()
filename = f"exceptions_{today}.xlsx"
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet.title = "Expiry in one month"
sheet.append(["SEP Ticket ID","Exception expiry date","L7_board_area","L6_board_area","L5_board_area","L4_board_area","# of accounts"])

for each_entry in exceptions_list:
    dict_values = list(each_entry.values())
    sheet.append(dict_values)
wb.save(filename)
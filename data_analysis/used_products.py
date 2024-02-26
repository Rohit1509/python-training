##### This script gives a dictionary of Used products and its total paybles

import openpyxl
import matplotlib.pyplot as plt

data_workbook = openpyxl.load_workbook('data.xlsx')
data_sheet = data_workbook["data"]

used_product_details = {}

for column in data_sheet.iter_cols():
    column_name = column[0].value
    if column_name == 'Product Type':
        for i, cell in enumerate(column):
            if i > 0 and cell.value == 'Used':
              used_product_details.update({data_sheet.cell(row=i+1, column=6).value:data_sheet.cell(row=i+1, column=15).value})

print(used_product_details)

def addlabels(product_names,product_values):
    for i in range(len(product_names)):
        plt.text(i,product_values[i],product_values[i])
        
product_names = list(used_product_details.keys())
product_values = list(used_product_details.values())

plt.bar(product_names, product_values)
addlabels(product_names, product_values)
plt.xlabel('Product Name that are USED')
plt.ylabel('Total Payble for the Product')
plt.title('Paybles for USED Products')
plt.show()
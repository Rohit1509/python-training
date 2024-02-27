##### This script gives a total number of NEW miners sold and total paybles from miners in 2022

import openpyxl

data_workbook = openpyxl.load_workbook('Invoices Data Analysis.xlsx')
data_sheet = data_workbook["Sheet1"]

total_miners_sold = {}

for column in data_sheet.iter_cols():
    column_name = column[0].value
    if column_name == 'Product Type':
        for i, cell in enumerate(column):
            date = str(data_sheet.cell(row=i+1, column=3).value).split(' ')[0]
            if i > 0 and cell.value == 'NEW' and data_sheet.cell(row=i+1, column=7).value == 'Miner' and date > '2022-01-01' and date < '2022-12-31':
              total_miners_sold.update({data_sheet.cell(row=i+1, column=6).value:data_sheet.cell(row=i+1, column=18).value})

print(total_miners_sold)

sum = 0

for each_value in list(total_miners_sold.values()):
    print(each_value)
    if each_value == 'None' or each_value == '1.08 BTC':
        continue
    else:
        sum = sum + each_value

total_number_miners_sold = len(list(total_miners_sold.keys()))
total_paybles_from_miners = sum

print(f'Total number of NEW miners sold in 2022 are {total_number_miners_sold}')
print(f'Total paybles received from NEW miners in 2022 are {total_paybles_from_miners}')


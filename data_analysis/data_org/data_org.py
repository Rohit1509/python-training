##### This script gives difference between incoming and outgoing payments

import openpyxl

data_workbook = openpyxl.load_workbook('data_org.xlsx')
incoming_data_sheet = data_workbook["Incoming Payments"]
outgoing_data_sheet = data_workbook["Outgoing Payments"]

incoming_payments = {}
outgoing_payments = {}
profit = {}

for column in incoming_data_sheet.iter_cols():
    column_name = column[0].value
    if column_name == 'Product Name':
        for i, cell in enumerate(column):
            if i > 0:
                incoming_payments.update({cell.value:incoming_data_sheet.cell(row=i+1, column=2).value})
                profit.update({cell.value:incoming_data_sheet.cell(row=i+1, column=4).value})                

print(incoming_payments)
print(profit)

for column in outgoing_data_sheet.iter_cols():
    column_name = column[0].value
    if column_name == 'Product Name':
        for i, cell in enumerate(column):
            if i > 0:
                outgoing_payments.update({cell.value:outgoing_data_sheet.cell(row=i+1, column=2).value})

print(outgoing_payments)

set1 = set(incoming_payments.items())
set2 = set(outgoing_payments.items())

difference_incoming_outgoing = { k : outgoing_payments[k] for k in set(outgoing_payments) - set(incoming_payments) }
print(difference_incoming_outgoing)
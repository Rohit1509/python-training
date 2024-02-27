##### This script gives a total number of pdu's sold and total paybles from pdu's

import openpyxl
import matplotlib.pyplot as plt

data_workbook = openpyxl.load_workbook('data.xlsx')
data_sheet = data_workbook["data"]

total_pdus_sold = {}

for column in data_sheet.iter_cols():
    column_name = column[0].value
    if column_name == 'Type of product':
        for i, cell in enumerate(column):
            if i > 0 and cell.value == 'Transformer':
              total_pdus_sold.update({data_sheet.cell(row=i+1, column=6).value:data_sheet.cell(row=i+1, column=16).value})

sum = 0

for each_value in list(total_pdus_sold.values()):
    print(each_value)
    if each_value == 'None' or each_value == '1.08 BTC':
        continue
    else:
        sum = sum + each_value

total_number_pdus_sold = len(list(total_pdus_sold.keys()))
total_paybles_from_pdus = sum

print(f'Total number of pdus sold are {total_number_pdus_sold}')
print(f'Total paybles received from pdus are {total_paybles_from_pdus}')


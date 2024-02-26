##### This script gives total number of unique invoices and total paybles.

import openpyxl

data_workbook = openpyxl.load_workbook('data.xlsx')
data_sheet = data_workbook["data"]

uniq_invoices = set()
paybles = []

for column in data_sheet.iter_cols():
    column_name = column[0].value
    if column_name == 'Order Number':
        for i, cell in enumerate(column):
            if i > 0:
              uniq_invoices.add(cell.value)
              paybles.append(data_sheet.cell(row=i+1, column=15).value)
              
print(f'The total number of unique invoices are {len(uniq_invoices)}')
print(f'Total sum of paybles is {sum(paybles)}')
##### This script gives a dictionary of Used products and its total paybles in 2022

import openpyxl
import time
import matplotlib.pyplot as plt

data_workbook = openpyxl.load_workbook('data.xlsx')
data_sheet = data_workbook["data"]

product_details_in_2022 = {}

for column in data_sheet.iter_cols():
    column_name = column[0].value
    if column_name == 'Invoice Date':
        for i, cell in enumerate(column):
            if i > 0:
              date = str(cell.value).split(' ')[0]
              if date > '2022-01-01' and date < '2022-12-31':
                  product_details_in_2022.update({data_sheet.cell(row=i+1, column=6).value:data_sheet.cell(row=i+1, column=15).value})
print(product_details_in_2022)

def addlabels(product_names,product_values):
    for i in range(len(product_names)):
        plt.text(i,product_values[i],product_values[i])
        
product_names = list(product_details_in_2022.keys())
product_values = list(product_details_in_2022.values())

plt.bar(product_names, product_values)
addlabels(product_names, product_values)
plt.xlabel('Product Name')
plt.ylabel('Total Payble for the Product')
plt.title('Paybles for Products sold in 2022')
plt.show()
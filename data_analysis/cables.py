##### This script gives a total number of cables sold and total paybles from cables

import openpyxl

data_workbook = openpyxl.load_workbook('Invoices Data Analysis.xlsx')
data_sheet = data_workbook["Sheet1"]

total_cables_sold = {}

for column in data_sheet.iter_cols():
    column_name = column[0].value
    if column_name == 'Type of product':
        for i, cell in enumerate(column):
            if i > 0 and cell.value == 'Cable':
              total_cables_sold.update({data_sheet.cell(row=i+1, column=6).value:data_sheet.cell(row=i+1, column=18).value})

sum = 0

for each_value in list(total_cables_sold.values()):
    print(each_value)
    if each_value == 'None' or each_value == '1.08 BTC':
        continue
    else:
        sum = sum + each_value

total_number_cables_sold = len(list(total_cables_sold.keys()))
total_paybles_from_cables = sum

print(f'Total number of cables sold are {total_number_cables_sold}')
print(f'Total paybles received from cables are {total_paybles_from_cables}')


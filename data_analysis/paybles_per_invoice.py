##### This script gives total number of paybles based on invoice number passed in the function.

import openpyxl

data_workbook = openpyxl.load_workbook('data.xlsx')
data_sheet = data_workbook["data"]

uniq_invoices = set()
paybles = []

def get_payble_per_invoice(invoice):
    for column in data_sheet.iter_cols():
        column_name = column[0].value
        if column_name == 'Order Number':
            for i, cell in enumerate(column):
                if i > 0 and cell.value == invoice:
                    paybles.append(data_sheet.cell(row=i+1, column=15).value)
    return paybles          

print(sum(get_payble_per_invoice(148)))
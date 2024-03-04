### Importing libraries

import re
import openpyxl # library for python to interact with excel
import datetime # library to get today's date
from dateutil.relativedelta import * # library to get date one month from now
import os # used to get path of the excel file on the PC

def prWarn(msg): print(f"\033[93mWarning: {msg}\033[00m")

def parse_date(dt,format=None):        
    if isinstance(dt, datetime.datetime):
        if format:
            return dt.strftime(format)
        else:
            return dt
    elif dt == "Permanent":
        return datetime.datetime.strptime(f"1/1/9999", "%d/%m/%Y")
    reg_exp = re.search("^Q(\d).(\d{4})$", dt)
    if reg_exp:
        quarter = int(reg_exp.group(1))
        year = int(reg_exp.group(2))
        if quarter == 4:
            year += 1
            month = 1
        else:
            month = quarter*3+1
        dt = datetime.datetime.strptime(f"1/{month}/{year}", "%d/%m/%Y")
        if format:
            return dt.strftime(format)
        return dt
    prWarn(f"Unable to parse date: {dt}")
    return datetime.datetime.strptime(f"1/1/1970", "%d/%m/%Y")

class Excel:
    def __init__(self, path): # Initialisation function to create coloumn mappings
        if not os.path.isfile(path):
            print(f"{path} does not exist")
            exit(1)
        self.path = path
        self.workbook = openpyxl.load_workbook(f"{path}") # read Excel file
        self.sheet = self.workbook["compliance"]
        self.col_mapping = {
            "serial_number": 0,
            "customer_name": 1,
            "compliance_document_name": 2,
            "validity": 3,
        }

    def get_exceptions_by_ticket(self, sep):
        validity_date = None
        name = None
        document_name = None
        document_serial_no = None
        for row in self.sheet[2:self.sheet.max_row]:
            for key in self.col_mapping.keys():
                exec(f'global {key}; {key} = row[self.col_mapping["{key}"]].value')
            if validity == None:
                continue
            expiry_date = parse_date(validity)
            todays_date = datetime.datetime.now()
            one_month_later_date = todays_date + relativedelta(months=+1)
            
            if expiry_date < one_month_later_date and serial_number == sep:
                document_serial_no = serial_number
                name = customer_name
                document_name = compliance_document_name
                validity_date = parse_date(validity,"%Y-%m-%d")
        return {'serial': document_serial_no, 'customer_name': name, 'compliance_document_name': document_name, 'expiry': validity_date}

path = os.path.join(os.getcwd(),"project.xlsx")

serial_numbers = set()
exceptions_list = []
exception_workbook = openpyxl.load_workbook(f"{path}")
exceptions_sheet = exception_workbook["compliance"]

for column in exceptions_sheet.iter_cols():
    column_name = column[0].value
    if column_name == 'Serial number':
        for i, cell in enumerate(column):
            if i > 0:
              serial_numbers.add(cell.value)
uniq_serial_nos = list(serial_numbers)


if __name__ == "__main__":
    if not os.path.isfile(path):
        print(f'Copy Orca_exceptions.xlsx to {os.path.join(os.getcwd(),"datamodel","data","SGS_Excel")}')
        exit(1)
    excel = Excel(path)
    for ticket in uniq_serial_nos:
      exception = excel.get_exceptions_by_ticket(ticket)
      exceptions_list.append(exception)

today = datetime.date.today()
filename = f"exceptions_{today}.xlsx"
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet.title = "Expiry in one month"
sheet.append(["Serial_no","Customer name","Compliance document name","expiry date"])

for each_entry in exceptions_list:
    dict_values = list(each_entry.values())
    sheet.append(dict_values)
wb.save(filename)